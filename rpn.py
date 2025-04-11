from flask import Flask, request, render_template, send_file
import pandas as pd
import os
import joblib
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads/processed/'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

MODEL_PATH = r"D:\Lakshya\Project\ICSS_VSCode\PROJECT\ICSS Web App\random_forest_model.pkl"

# Load Model Function
def load_model():
    try:
        model_data = joblib.load(MODEL_PATH)
        model, vectorizer = model_data if isinstance(model_data, tuple) else (model_data, None)
        if not hasattr(model, 'predict'):
            raise ValueError("Loaded model does not have a 'predict' method")
        print("Model loaded successfully")
        return model, vectorizer
    except Exception as e:
        print(f"Error loading model: {e}")
        return None, None

model, vectorizer = load_model()

# Function to Calculate RPN
def calculate_rpn(severity, occurrence, detection):
    return severity * occurrence * detection

# Function to Predict Priority & RPN
def predict_priority(observation):
    try:
        if model and vectorizer:
            transformed_text = vectorizer.transform([str(observation)])
            priority = model.predict(transformed_text)[0]
            
            severity_map = {'High': 10, 'Moderate': 5, 'Low': 2}
            occurrence_map = {'High': 9, 'Moderate': 6, 'Low': 3}
            detection_map = {'High': 2, 'Moderate': 5, 'Low': 8}
            
            severity = severity_map.get(priority, 1)
            occurrence = occurrence_map.get(priority, 5)
            detection = detection_map.get(priority, 4)
            
            rpn = calculate_rpn(severity, occurrence, detection)
            return priority, rpn
    except Exception as e:
        print(f"Prediction Error: {e}")
    return 'Low', 10

# Retrain Model Function
def retrain_model(filepath):
    global model, vectorizer
    try:
        df = pd.read_excel(filepath)
        if 'Observation' not in df.columns or 'Priority' not in df.columns:
            print("Missing required columns for training")
            return
        
        X, y = df['Observation'].astype(str), df['Priority']
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
        
        vectorizer = TfidfVectorizer()
        X_train_tfidf = vectorizer.fit_transform(X_train)
        X_test_tfidf = vectorizer.transform(X_test)
        
        model = RandomForestClassifier(n_estimators=100, random_state=42)
        model.fit(X_train_tfidf, y_train)
        
        y_pred = model.predict(X_test_tfidf)
        accuracy = accuracy_score(y_test, y_pred)
        print(f"New Model Accuracy: {accuracy * 100:.2f}%")
        
        joblib.dump((model, vectorizer), MODEL_PATH)
        print("Model retrained and updated.")
        model, vectorizer = load_model()
    except Exception as e:
        print(f"Error in model retraining: {e}")

@app.route('/')
def index():
    return render_template('front1.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'complaint_file' not in request.files:
        return "No complaint_file part", 400
    
    file = request.files['complaint_file']
    if file.filename == '':
        return "No selected file", 400
    
    if file:
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        
        try:
            df = pd.read_excel(filepath)
        except Exception as e:
            return f"Error reading file: {e}", 400
        
        if 'Observation' not in df.columns or 'Incident Status' not in df.columns:
            return "Error: Required columns are missing in the uploaded file.", 400
        
        retrain_model(filepath)

        # Assign Priority and RPN correctly
        df[['Priority', 'RPN']] = df['Observation'].fillna('').apply(lambda obs: pd.Series(predict_priority(obs)))

        # Sorting based on priority order
        priority_order = {'High': 3, 'Moderate': 2, 'Low': 1}
        df['Priority_Value'] = df['Priority'].map(priority_order)
        df = df.sort_values(by='Priority_Value', ascending=False).drop(columns=['Priority_Value'])

        processed_filepath = os.path.join(UPLOAD_FOLDER, 'processed_' + file.filename)
        df.to_excel(processed_filepath, index=False)

        # Apply colors in Excel sheet
        wb = load_workbook(processed_filepath)
        ws = wb.active
        
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    status = cell.value.lower()
                    if "closed" in status or "completed" in status:
                        cell.fill = green_fill
                    elif "pending" in status:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill

        # **Ensure RPN is saved properly in the Excel file**
        wb.save(processed_filepath)

        return send_file(processed_filepath, as_attachment=True)


        # Assign Priority and RPN separately
        df['Priority'] = df['Observation'].fillna('').apply(lambda obs: predict_priority(obs)[0])
        df['RPN'] = df['Observation'].fillna('').apply(lambda obs: predict_priority(obs)[1])
        
        # Sorting based on priority order
        priority_order = {'High': 3, 'Moderate': 2, 'Low': 1}
        df['Priority_Value'] = df['Priority'].map(priority_order)
        df = df.sort_values(by='Priority_Value', ascending=False).drop(columns=['Priority_Value'])
        
        processed_filepath = os.path.join(UPLOAD_FOLDER, 'processed_' + file.filename)
        df.to_excel(processed_filepath, index=False)
        
        wb = load_workbook(processed_filepath)
        ws = wb.active
        
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    status = cell.value.lower()
                    if "closed" in status or "completed" in status:
                        cell.fill = green_fill
                    elif "pending" in status:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill
        df['RPN'] = df['Observation'].fillna('').apply(lambda obs: predict_priority(obs)[1])
        wb.save(processed_filepath)
        return send_file(processed_filepath, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
