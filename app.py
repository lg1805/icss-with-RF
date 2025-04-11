from flask import Flask, request, render_template, send_file
import pandas as pd
import os
import joblib
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads/processed/'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

MODEL_PATH = r"D:\KOEL\ICSS\Deployment - Copy\icss-backend\model\random_forest_model.pkl"

# ✅ Function to Load Model
def load_model():
    try:
        model_data = joblib.load(MODEL_PATH)
        if isinstance(model_data, tuple):
            model, vectorizer = model_data
        else:
            model = model_data
            vectorizer = None
        
        if not hasattr(model, 'predict'):
            raise ValueError("Loaded model does not have a 'predict' method")
        print("Model loaded successfully")
        return model, vectorizer
    except Exception as e:
        print(f"Error loading model: {e}")
        return None, None

model, vectorizer = load_model()

# ✅ Function to Retrain the Model on New Data
def retrain_model(filepath):
    global model, vectorizer  # Declare them as global before usage
    try:
        df = pd.read_excel(filepath)
        if 'Observation' not in df.columns or 'Priority' not in df.columns:
            print("Missing required columns for training")
            return

        X, y = df['Observation'].astype(str), df['Priority']
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

        vectorizer = TfidfVectorizer()
        X_train_tfidf = vectorizer.fit_transform(X_train)
        X_test_tfidf = vectorizer.transform(X_test)

        model = RandomForestClassifier(n_estimators=100, random_state=42)
        model.fit(X_train_tfidf, y_train)

        # Evaluate Model Performance
        y_pred = model.predict(X_test_tfidf)
        accuracy = accuracy_score(y_test, y_pred)
        print(f"New Model Accuracy: {accuracy * 100:.2f}%")

        # Save Updated Model
        joblib.dump((model, vectorizer), MODEL_PATH)
        print("Model retrained and updated.")

        # Reload the new model for use
        model, vectorizer = load_model()
    except Exception as e:
        print(f"Error in model retraining: {e}")


@app.route('/')
def index():
    return render_template('front1.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'complaint_file' not in request.files:
        return "No complaint_file part", 400
    
    file = request.files['complaint_file']
    if file.filename == '':
        return "No selected file", 400
    
    if file:
        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)

        try:
            df = pd.read_excel(filepath)
        except Exception as e:
            return f"Error reading file: {e}", 400
        
        if 'Observation' not in df.columns or 'Incident Status' not in df.columns:
            return "Error: Required columns are missing in the uploaded file.", 400
        
        # 🚀 **Step 1: Retrain Model Automatically**
        retrain_model(filepath)

        # 🚀 **Step 2: Segregate SPN and Non-SPN complaints**
        spn_df = df[df['Observation'].str.contains('SPN', case=False, na=False)]
        non_spn_df = df[~df['Observation'].str.contains('SPN', case=False, na=False)].copy()

        # 🚀 **Step 3: Assign Priority Using Updated Model**
        def predict_priority(observation):
            try:
                if model and vectorizer:
                    transformed_text = vectorizer.transform([str(observation)])
                    return model.predict(transformed_text)[0]
            except Exception as e:
                print(f"Prediction Error: {e}")
            return 'Low'
        
        if not non_spn_df.empty:
            non_spn_df['Priority'] = non_spn_df['Observation'].fillna('').apply(predict_priority)
        
        # 🚀 **Step 4: Sort by Priority**
        priority_order = {'High': 3, 'Moderate': 2, 'Low': 1}
        non_spn_df['Priority_Value'] = non_spn_df['Priority'].map(priority_order)
        non_spn_df = non_spn_df.sort_values(by='Priority_Value', ascending=False).drop(columns=['Priority_Value'])
        
        # 🚀 **Step 5: Save Processed File**
        processed_filepath = os.path.join(UPLOAD_FOLDER, 'processed_' + file.filename)
        with pd.ExcelWriter(processed_filepath, engine='openpyxl') as writer:
            spn_df.to_excel(writer, sheet_name='SPN_Complaints', index=False)
            non_spn_df.to_excel(writer, sheet_name='Non_SPN_Complaints', index=False)
        
        # 🚀 **Step 6: Apply Formatting**
        wb = load_workbook(processed_filepath)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            incident_status_col = None
            for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
                if col[0].value == 'Incident Status':
                    incident_status_col = col_idx
                    break
            
            if incident_status_col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=incident_status_col)
                    if isinstance(cell.value, str):
                        status_lower = cell.value.lower()
                        if "closed" in status_lower or "completed" in status_lower:
                            cell.fill = green_fill
                        elif "pending" in status_lower:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill
        
        wb.save(processed_filepath)
        return send_file(processed_filepath, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
